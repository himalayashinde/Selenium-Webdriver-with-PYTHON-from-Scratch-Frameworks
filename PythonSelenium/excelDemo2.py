import openpyxl

book = openpyxl.load_workbook("I:\\Coding_journey\\Selenium-Webdriver-with-PYTHON-from-Scratch-Frameworks\\PythonSelenium\\PythonDemo.xlsx")

sheet = book.active
dictionary={}
cell= sheet.cell(row=1,column=2)

print(cell.value)

sheet.cell(row=2,column=2).value = "Himalaya"
print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

print("############# FOR LOOP IS BELOW ###############")

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase5":
        for j in range(1,sheet.max_column+1):
            dictionary[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
            # print(sheet.cell(row=i,column=j).value)

print(dictionary)

# Save the workbook to persist the changes
# book.save("I:\\Coding_journey\\Selenium-Webdriver-with-PYTHON-from-Scratch-Frameworks\\PythonSelenium\\PythonDemo.xlsx")
