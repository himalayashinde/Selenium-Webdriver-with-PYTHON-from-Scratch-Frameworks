import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

file_path="C:\\Users\\ADMIN\\Downloads\\download.xlsx"

def update_excel_data(file_path, searchTerm, colName, new_val):
    book = openpyxl.load_workbook(file_path)
    sheet= book.active
    Dict={}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"]=i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Dict["row"]=i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_val
    book.save(file_path)




fruit_name="Apple"
newValue= "999"
columnName="price"

base_dir = os.path.dirname(__file__)
print(base_dir)

# Construct path to the ChromeDriver inside the resources folder
driver_path = os.path.join(base_dir, "chromedriver-win64", "chromedriver.exe")

service_obj = Service(driver_path)
driver= webdriver.Chrome(service=service_obj)

driver.implicitly_wait(5)
# driver = webdriver.Chrome()
driver.maximize_window()

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

# download the excel
driver.find_element(By.ID,"downloadButton").click()

# Edit the excel with updated values

update_excel_data(file_path, fruit_name, columnName, newValue)


# upload th excel
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)
toast_locator=(By.XPATH,"//div[text()='Updated Excel Data Successfully.']")
wait=WebDriverWait(driver,5)
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)


price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price= driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text

print(actual_price)

assert actual_price == newValue